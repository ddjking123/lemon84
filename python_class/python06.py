# 123
"""
接口自动化测试
1、测试数据从excel用例里读取出来--read_date() --
2、发送接口请求，得到响应消息--执行结果--函数--api_request
3、执行结果 vs 预期结果对比  -相同通过；不同就不通过--断言===结论-pass fail
4、把最终的结果写入excel中，写入数据一定要保存才能生效 --write_result
5、把最终的结果写入excel表格中
操作Excel表格导入数据，openpyxl、
1、安装：pip/pycharm
2、导入--  import ：导入所有
from xxx import xxx:导入局部

主要操作：
1、加载工作簿对象：
2、表单对象：sheet
3、单元格：行列
4、写入数据到单元格中---保存才会生效
wb = load_workbook('testcase_api_wuye.xlsx')  #加载工作簿
sheet = wb['register'] #选择表单
cell = sheet.cell(row=1,column=1)  #单元格
print(cell.value)
cell.value = '用例编号'  #重新赋值 -写入数据
print(cell.value)
wb.save('testcase_api_wuye.xlsx')  #保存文件名，也可以改名字就是另存为，要关闭打开的excel

把一个功能封装成函数步骤：
1、def 函数名
2、参数化--变量值
3、是否需要定义返回值 --有没有东西给人使用
"""
import requests
from openpyxl import load_workbook



# 接口请求函数
def api_request(url,data):
    head = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}
    response = requests.post(url=url,json=data,headers=head)   #形参url、data待传入
    return response.json()

# from openpyxl import load_workbook
# import openpyxl
# wb = load_workbook('testcase_api_wuye.xlsx')  #加载工作铺
# sheet = wb['register']   #加载表单
# cell = sheet.cell(row =1,column =1)
# print(cell.value)

#上传接口
def api_uplod(url,head,file):
    url = 'http://47.115.15.198:7001/smarthome/file/upload'
    head1 = {'X-Lemonban-Media-Type':'lemonban.v2'}
    # 上传文件参数
    file = {'file': ('test.png',open('test.png','rb'),'image/png')}
    # 发送请求
    respinse_u = requests.post(url=url,files=file,headers=head1)
    # 打印结果
    print(respinse_u.json())

# 读取测试用例数据

# cases = []   #空列表，用于储存所有数据
# wb = load_workbook('testcase_api_wuye.xlsx')  #加载工作簿
# sheet = wb['register'] #选择表单
# max_r = sheet.max_row   #获取最大行号
# for i in range(2,max_r+1):
#     case = dict(
#     case_id =sheet.cell(row=i,column=1).value,
#     url = sheet.cell(row=i,column=5).value, #获取url地址
#     data = sheet.cell(row=i,column=6).value, #获取参数
#     expected_reslut = sheet.cell(row=i,column=7).value  #获取期望
#     )
#     cases.append(case)
# print(cases)

# 参数化
def read_data(file_name,sheet_name):
    cases = []   #空列表，用于储存所有数据
    wb = load_workbook(file_name)  #加载工作簿
    sheet = wb[sheet_name] #选择表单
    max_r = sheet.max_row   #获取最大行号
    for i in range(2,max_r+1):
        case = dict(
        case_id =sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value, #获取url地址
        data = sheet.cell(row=i,column=6).value, #获取参数
        expected_reslut = sheet.cell(row=i,column=7).value  #获取期望
        )
        cases.append(case)
    return (cases)  #返回所有用例

#写入结果的函数
def write_result(filename,sheetname,row,cloumn,final_result):
    wb = load_workbook(filename)  #加载工作铺
    sheet = wb[sheetname]   #加载表单
    cell = sheet.cell(row =row,column =cloumn).value = final_result#单元格
    wb.save(filename)

# eval():内置函数 === 被引号包裹的-Python表达式取出来==去掉Python表达式中的引号

# cases = read_data('testcase_api_wuye.xlsx','register')
# for case in cases:
#     case_id = case.get('case_id')
#     url = case.get('url')  #获取url
#     data =case['data']   #字典的另一种取值法，获取参数,excel中取到的是字符串，需要转化为字典
#     data = eval(data)  #字符串--转换成字典
#     expected_result = case.get("expected_reslut")#获取期望结果
#     expected_result = eval(expected_result)
#     real_result = api_request(url=url,data=data)  #调用 发送按接口请求函数--传参
#     real_msg = real_result.get('msg')  #执行结果的msg
#     expected_msg = expected_result['msg']  #预期结果的msg
#     print('执行结果:{}'.format(expected_msg))
#     print('期望结果：{}'.format(real_msg))
#     if real_result==expected_result:
#         print('第{}条用例通过'.format(case_id))
#         final = 'Passed'
#     else:
#         print('第{}条用例不通过'.format(case_id))
#         final = 'Failed'
#     print('*'*10)
#     write_result('testcase_api_wuye.xlsx','register',case_id+1,8,final)

#参数化
def excute(filename,sheetname):
    cases = read_data(filename, sheetname)
    for case in cases:
        case_id = case.get('case_id')
        url = case.get('url')  # 获取url
        data = case['data']  # 字典的另一种取值法，获取参数,excel中取到的是字符串，需要转化为字典
        data = eval(data)  # 字符串--转换成字典
        expected_result = case.get("expected_reslut")  # 获取期望结果
        expected_result = eval(expected_result)
        real_result = api_request(url=url, data=data)  # 调用 发送按接口请求函数--传参
        real_msg = real_result.get('msg')  # 执行结果的msg
        expected_msg = expected_result['msg']  # 预期结果的msg
        print('执行结果:{}'.format(expected_msg))
        print('期望结果：{}'.format(real_msg))
        if real_result == expected_result:
            print('第{}条用例通过'.format(case_id))
            final = 'Passed'
        else:
            print('第{}条用例不通过'.format(case_id))
            final = 'Failed'
        print('*' * 10)
        write_result(filename,sheetname,case_id+1,8,final)  #调用写入结果函数--传参
excute('testcase_api_wuye.xlsx', 'login')